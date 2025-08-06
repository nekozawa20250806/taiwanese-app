import streamlit as st
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter
import io
import json
import openai
import os

openai.api_key = os.getenv("OPENAI_API_KEY")

st.title("台湾華語文章解析ツール（Excel & CSV出力）")

# モデル選択UI
model_choice = st.selectbox("モデルを選択してください", ["GPT-4", "GPT-3.5"], index=0)
if model_choice == "GPT-4":
    model_name = "gpt-4"
else:
    model_name = "gpt-3.5-turbo"

st.markdown("### 台湾華語の文章を入力してください（1行1文）：")
input_text = st.text_area("文章入力", height=200, placeholder="隨著國家經濟發展至一定程度...")

if st.button("解析してファイル生成"):
    if not input_text.strip():
        st.warning("文章を入力してください")
    else:
        sentences = input_text.strip().split("\n")

        sentences_data = []
        glossary_data = []

        st.info(f"{model_choice} で解析中...（{len(sentences)} 文）")

        for sentence in sentences:
            prompt = f"""
以下の台湾華語の文章を解析し、JSON形式で返してください。

文章: {sentence}

出力形式：
{{
"japanese": "自然な日本語訳（要点補足付き）",
"pinyin": "声調記号付きピンイン",
"glossary": [
  {{"term": "台湾華語単語", "meaning": "日本語の意味", "pinyin": "声調記号付き"}}
]
}}
"""
            try:
                response = openai.ChatCompletion.create(
                    model=model_name,
                    messages=[{"role": "user", "content": prompt}],
                    temperature=0.7
                )
                result = response.choices[0].message["content"]
                parsed = json.loads(result)

                sentences_data.append([sentence, parsed["japanese"], parsed["pinyin"]])
                for g in parsed["glossary"]:
                    glossary_data.append([g["term"], g["meaning"], g["pinyin"]])

            except Exception as e:
                st.error(f"エラーが発生しました: {e}")
                st.stop()

        df_sentences = pd.DataFrame(sentences_data, columns=["台湾華語", "日本語訳", "ピンイン"])
        df_glossary = pd.DataFrame(glossary_data, columns=["台湾華語", "日本語訳", "ピンイン"])

        # Excel作成
        excel_buffer = io.BytesIO()
        wb = Workbook()
        ws1 = wb.active
        ws1.title = "文章"
        ws1.append(["台湾華語", "日本語訳", "ピンイン"])
        for r in df_sentences.itertuples(index=False):
            ws1.append(r)

        ws2 = wb.create_sheet("語釈")
        ws2.append(["台湾華語", "日本語訳", "ピンイン"])
        for r in df_glossary.itertuples(index=False):
            ws2.append(r)

        for ws in [ws1, ws2]:
            ws.freeze_panes = "A2"
            for col in range(1, 4):
                ws.column_dimensions[get_column_letter(col)].width = 50
            for row in ws.iter_rows():
                max_height = 18
                for cell in row:
                    cell.alignment = Alignment(wrap_text=True, vertical="top")
                    cell.font = Font(name="Noto Sans CJK JP Medium", size=11)
                    if cell.value:
                        lines = str(cell.value).count("\n") + 1
                        est_lines = max(lines, len(str(cell.value)) // 50 + 1)
                        height = est_lines * 18
                        if height > max_height:
                            max_height = height
                ws.row_dimensions[row[0].row].height = max_height

        wb.save(excel_buffer)
        excel_buffer.seek(0)

        csv_sent = df_sentences.to_csv(index=False, encoding="utf-8-sig")
        csv_gloss = df_glossary.to_csv(index=False, encoding="utf-8-sig")

        st.success("ファイル生成完了！以下からダウンロードできます：")
        st.download_button("📥 Excelファイル", excel_buffer, file_name="taiwanese_analysis.xlsx")
        st.download_button("📥 文章CSV", csv_sent, file_name="taiwanese_sentences.csv")
        st.download_button("📥 語釈CSV", csv_gloss, file_name="taiwanese_glossary.csv")
